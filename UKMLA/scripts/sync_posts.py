import os
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, '..', 'scratch', 'parsed_posts.json')
    excel_path = os.path.join(script_dir, '..', 'UKMLA_Posts_Registry.xlsx')
    
    if not os.path.exists(json_path):
        print(f"Error: JSON file not found at {json_path}")
        return

    with open(json_path, 'r', encoding='utf-8') as f:
        posts = json.load(f)

    # Convert to DataFrame
    df = pd.DataFrame(posts)

    # Rename columns to match exact request and make them look neat
    column_mapping = {
        'post_type': 'post_type',
        'post_title': 'post_title',
        'post_slug': 'post_slug',
        'post_excerpt': 'post_excerpt',
        'category': 'category',
        'tags': 'tags',
        'source_full_url': 'source_full_url',
        'seo_title': 'seo_title',
        'seo_description': 'seo_description',
        'primary_keyword': 'Primary keyword',
        'internal_link': 'Internal Link',
        'html_content': 'html_content',
        'featured_image_keyword': 'featured_image_keyword',
        'featured_image_url': 'featured_image_url',
        'featured_image_title': 'featured_image_title',
        'feature_image_alt_text': 'feature_image_alt_text',
        'word_count': 'Word Count',
        'outgoing_links': 'Outgoing Links',
        'incoming_links': 'Incoming Links',
        'weaving_reference': 'Weaving Reference'
    }

    df = df.rename(columns=column_mapping)

    # Ensure all columns are present in this order
    columns_order = [
        'post_type', 'post_title', 'post_slug', 'post_excerpt', 'category', 'tags',
        'source_full_url', 'seo_title', 'seo_description', 'Primary keyword', 'Internal Link',
        'html_content', 'featured_image_keyword', 'featured_image_url', 'featured_image_title',
        'feature_image_alt_text', 'Word Count', 'Outgoing Links', 'Incoming Links', 'Weaving Reference'
    ]

    df = df[columns_order]

    # Write to Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Posts Registry', index=False)
        
        # Also create a quick info / Read Me sheet for weaving reference
        readme_data = {
            'Section': [
                'Purpose', 
                'Weaving Concept', 
                'Target Audience', 
                'Excel Schema',
                'How to use this registry'
            ],
            'Details': [
                'Tracks all posts published on the website along with their SEO metadata, keyword focus, and images.',
                'The Columns "Incoming Links" and "Outgoing Links" record the internal relationships between articles to ensure a strong Topic Cluster and semantic web.',
                'Global medical students and International Medical Graduates (IMGs). Content is written without local bias but targets regional search trends (like India, Pakistan, Nigeria) naturally.',
                'Primary columns: post_type, post_title, post_slug, post_excerpt, category, tags, source_full_url, seo_title, seo_description, Primary keyword, Internal Link, html_content, featured_image_keyword, featured_image_url (Higgsfield AI), featured_image_title, feature_image_alt_text.',
                '1. Generate a new post using the "UKMLA post" trigger. 2. Append the post object in News.jsx. 3. Run the sync script to update this Excel sheet and re-verify the internal links (weaving).'
            ]
        }
        readme_df = pd.DataFrame(readme_data)
        readme_df.to_excel(writer, sheet_name='Read Me & Weaving Info', index=False)

    # Open sheet to style it
    wb = openpyxl.load_workbook(excel_path)
    
    # 1. Style Read Me & Weaving Info sheet
    ws_readme = wb['Read Me & Weaving Info']
    ws_readme.column_dimensions['A'].width = 30
    ws_readme.column_dimensions['B'].width = 80
    
    # Style title header of Read Me
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    for cell in ws_readme[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    for row in range(2, ws_readme.max_row + 1):
        ws_readme.cell(row=row, column=1).font = Font(name="Segoe UI", size=11, bold=True)
        ws_readme.cell(row=row, column=2).font = Font(name="Segoe UI", size=10)
        ws_readme.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        
    ws_readme.row_dimensions[1].height = 25

    # 2. Style Posts Registry sheet
    ws_reg = wb['Posts Registry']
    
    # Style header row
    reg_header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid") # Teal theme
    reg_header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for col_idx, cell in enumerate(ws_reg[1], 1):
        cell.fill = reg_header_fill
        cell.font = reg_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    ws_reg.row_dimensions[1].height = 28
    
    # Enable wrapping and set column widths
    column_widths = {
        'A': 15,  # post_type
        'B': 35,  # post_title
        'C': 30,  # post_slug
        'D': 40,  # post_excerpt
        'E': 15,  # category
        'F': 25,  # tags
        'G': 40,  # source_full_url
        'H': 35,  # seo_title
        'I': 45,  # seo_description
        'J': 22,  # Primary keyword
        'K': 20,  # Internal Link
        'L': 50,  # html_content
        'M': 22,  # featured_image_keyword
        'N': 40,  # featured_image_url
        'O': 25,  # featured_image_title
        'P': 30,  # feature_image_alt_text
        'Q': 12,  # Word Count
        'R': 25,  # Outgoing Links
        'S': 25,  # Incoming Links
        'T': 30   # Weaving Reference
    }
    
    for col_letter, width in column_widths.items():
        ws_reg.column_dimensions[col_letter].width = width
        
    # Style rows
    for row in range(2, ws_reg.max_row + 1):
        ws_reg.row_dimensions[row].height = 24
        for col in range(1, len(columns_order) + 1):
            cell = ws_reg.cell(row=row, column=col)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            
            # Text alignment
            col_letter = get_column_letter(col)
            if col_letter in ['A', 'E', 'K', 'Q']:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_letter in ['L', 'D', 'I']:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(excel_path)
    print(f"Successfully generated and styled Excel registry at {excel_path}")

if __name__ == "__main__":
    main()
