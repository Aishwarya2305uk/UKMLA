# -*- coding: utf-8 -*-
import json
import openpyxl
from openpyxl.styles import Font
from copy import copy

FILE = 'UKMLA_SEO_Keyword_Research (1).xlsx'

wb = openpyxl.load_workbook(FILE)

# ---------- 1. Move the 6 stale "All Keywords - Unique" rows to "Removed (On Website)" ----------
ws_unique = wb['All Keywords - Unique']
ws_removed = wb['Removed (On Website)']

# batch-4 slug that covers each of the 6 keywords, in the same row order as the sheet
covered_by = {
    'how to pass UKMLA on first attempt': 'how-to-pass-ukmla-on-first-attempt',
    'UKMLA syllabus 2026 changes': 'ukmla-syllabus-2026-changes',
    'foundation year vs specialty training IMG': 'foundation-year-vs-specialty-training-for-imgs',
    'GMC registration refund policy': 'gmc-registration-refund-policy',
    'UKMLA Anki deck download': 'ukmla-anki-deck-and-flashcard-revision-guide',
    'UKMLA crash course for doctors': 'ukmla-crash-course-for-doctors',
}

# read the 6 data rows (rows 4-9: header on row 3, data rows 4-9)
unique_rows = []
for r in range(4, 10):
    vals = [ws_unique.cell(row=r, column=c).value for c in range(1, 12)]
    if vals[1]:
        unique_rows.append(vals)

# append to Removed (On Website): columns are #, Keyword, Country/Region, Est. Monthly Searches, Already Covered By
last_removed_row = ws_removed.max_row
next_num = ws_removed.cell(row=last_removed_row, column=1).value + 1
write_row = last_removed_row + 1
for vals in unique_rows:
    keyword = vals[1]
    country = vals[3]
    searches = vals[5]
    slug = covered_by.get(keyword, '')
    ws_removed.cell(row=write_row, column=1, value=next_num)
    ws_removed.cell(row=write_row, column=2, value=keyword)
    ws_removed.cell(row=write_row, column=3, value=country)
    ws_removed.cell(row=write_row, column=4, value=searches)
    ws_removed.cell(row=write_row, column=5, value=f'/news#{slug}')
    next_num += 1
    write_row += 1

# clear the 6 data rows in All Keywords - Unique (rows 4-9)
for r in range(4, 10):
    for c in range(1, 12):
        ws_unique.cell(row=r, column=c, value=None)

# update the summary header rows (row 1 title, row 2 note)
ws_unique.cell(row=1, column=1, value='All UKMLA Keywords - Unique & Deduplicated (0 keywords)')
ws_unique.cell(row=2, column=1, value=(
    'Cleaned on 2 July 2026: the final 6 genuine gaps identified on 2026-07-02 were written up as posts '
    '(batch of 6, see Post Registry) and moved to the Removed (On Website) tab. No outstanding keyword-sheet '
    'gaps remain. The two subsequent batches (20 posts, then 50 posts) were sourced from live UKMLA/GMC/NHS '
    'news research and manual re-mining of the audience-specific tabs (Medical Students / Doctors / India / '
    'Other Countries Keywords) rather than this master list — check those tabs plus fresh regulatory news '
    'before starting a future batch.'
))

# ---------- 2. Refresh the Post Registry tab with all 157 live posts ----------
ws_reg = wb['Post Registry']
rows = json.load(open('scratch/post_registry_rows.json', encoding='utf-8'))

# unmerge the old totals-footer row (A55:E55, H55:I55) before clearing/rewriting
for rng in list(ws_reg.merged_cells.ranges):
    if str(rng) in ('A55:E55', 'H55:I55'):
        ws_reg.unmerge_cells(str(rng))

# clear existing data + footer rows (row 4 onward; row 3 is header)
if ws_reg.max_row >= 4:
    for r in range(4, ws_reg.max_row + 1):
        for c in range(1, 10):
            ws_reg.cell(row=r, column=c, value=None)

def status_for(words):
    if words >= 2500:
        return 'Full'
    if words >= 1000:
        return 'Amber'
    return 'Red'

r = 4
full_count = 0
amber_count = 0
red_count = 0
for i, row in enumerate(rows, start=1):
    status = status_for(row['words'])
    if status == 'Full':
        full_count += 1
    elif status == 'Amber':
        amber_count += 1
    else:
        red_count += 1
    ws_reg.cell(row=r, column=1, value=i)
    ws_reg.cell(row=r, column=2, value=row['slug'])
    ws_reg.cell(row=r, column=3, value=row['title'])
    ws_reg.cell(row=r, column=4, value=row['tag'])
    ws_reg.cell(row=r, column=5, value=row['date'])
    ws_reg.cell(row=r, column=6, value=row['words'])
    ws_reg.cell(row=r, column=7, value=row['chars'])
    ws_reg.cell(row=r, column=8, value=status)
    ws_reg.cell(row=r, column=9, value=None)
    r += 1

footer_row = r
ws_reg.cell(row=footer_row, column=1, value=f'TOTALS  ({len(rows)} posts)')
ws_reg.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=5)
status_summary = f'Full: {full_count}'
if amber_count:
    status_summary += f'  |  Amber: {amber_count}'
if red_count:
    status_summary += f'  |  Stub: {red_count}'
ws_reg.cell(row=footer_row, column=8, value=status_summary)
ws_reg.merge_cells(start_row=footer_row, start_column=8, end_row=footer_row, end_column=9)

ws_reg.cell(row=1, column=1, value=f'UKMLA Blog Post Registry  ·  {len(rows)} Posts  ·  Last Updated 2 July 2026')

wb.save(FILE)
print('saved', FILE)
print('Post Registry rows written:', len(rows))
print('All Keywords - Unique: 6 rows moved to Removed (On Website)')
