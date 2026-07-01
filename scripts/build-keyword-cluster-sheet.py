import json
import re
from collections import defaultdict

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

STOPWORDS = {
    'ukmla', 'the', 'a', 'an', 'for', 'and', 'in', 'of', 'to', 'is', 'what', 'how',
    'your', 'you', 'on', 'with', 'are', 'it', 'this', 'that', 'or', 'be', 'do',
}


def tokenize(text):
    return {w for w in re.findall(r"[a-z0-9]+", (text or '').lower()) if w not in STOPWORDS and len(w) > 2}


# ── Load source data ─────────────────────────────────────────────────────────
with open('scratch/posts_min.json', encoding='utf-8') as f:
    posts = json.load(f)

with open('scripts/linkage-data.json', encoding='utf-8') as f:
    linkage = json.load(f)

link_by_key = {r['key']: r for r in linkage}
pillar_title_by_url = {r['key']: r['title'] for r in linkage if r['type'] == 'pillar'}

kw_wb = openpyxl.load_workbook('UKMLA_SEO_Keyword_Research.xlsx')

# "All Keywords - Unique" — header row is row 3 (rows 1-2 are title/subtitle)
ws = kw_wb['All Keywords - Unique']
kw_rows = []
for i, r in enumerate(ws.iter_rows(min_row=3, values_only=True)):
    if i == 0:
        headers = r
        continue
    if r[1] is None:
        continue
    kw_rows.append(dict(zip(headers, r)))

title_to_intent = {row['Keyword / Search Phrase']: row.get('Search Intent') for row in kw_rows}

# "Removed (On Website)" — real researched keywords already mapped to a pillar page
ws2 = kw_wb['Removed (On Website)']
pillar_keywords = defaultdict(list)  # url -> [keyword, ...] ordered by search volume desc
pillar_kw_rows = []
for i, r in enumerate(ws2.iter_rows(min_row=4, values_only=True)):
    if i == 0:
        continue  # header row itself
    if not r or not isinstance(r[1], str):
        continue
    kw_text, searches, covered_by = r[1], r[3], r[4]
    if not covered_by:
        continue
    m = re.search(r'\(([^)]+)\)', str(covered_by))
    if not m:
        continue
    url = m.group(1)
    pillar_kw_rows.append((url, kw_text, searches or 0))

for url, kw_text, searches in sorted(pillar_kw_rows, key=lambda x: -x[2]):
    pillar_keywords[url].append(kw_text)

# "Keyword Clusters" hub sheet — used to infer Search Intent for pillar pages
ws3 = kw_wb['Keyword Clusters']
hub_rows = []
for i, r in enumerate(ws3.iter_rows(min_row=4, values_only=True)):
    if i == 0:
        continue
    if not r or not r[1]:
        continue
    hub_rows.append({'pillar_term': r[1], 'intent': r[3]})


def infer_pillar_intent(pillar_title):
    tokens = tokenize(pillar_title)
    best, best_score = 'Informational', 0
    for hub in hub_rows:
        score = len(tokens & tokenize(hub['pillar_term']))
        if score > best_score:
            best_score, best = score, hub['intent']
    return best or 'Informational'


# ── Secondary keyword matching for blog posts ────────────────────────────────
def secondary_keywords_for_post(post, n=5):
    # Weight the primary keyword's own tokens double — title alone is too broad
    # and, combined with `tag` (shared by up to 22 posts in one cluster), used
    # to produce near-identical secondary-keyword lists across a whole cluster.
    pk_tokens = tokenize(post.get('primaryKeyword', ''))
    title_tokens = tokenize(post['title'])
    query_tokens = pk_tokens | title_tokens
    own_title = post['title']
    own_pk = (post.get('primaryKeyword') or '')
    scored = []
    for row in kw_rows:
        text = row['Keyword / Search Phrase']
        if text == own_title or text.lower() == own_pk.lower():
            continue
        toks = tokenize(text)
        if not toks:
            continue
        pk_overlap = len(pk_tokens & toks)
        total_overlap = len(query_tokens & toks)
        if total_overlap == 0:
            continue
        weighted_overlap = total_overlap + pk_overlap  # primary-keyword tokens count double
        score = weighted_overlap / max(len(toks), 1)
        has_volume = 1 if row.get('Est. Monthly Searches') else 0
        scored.append((score, has_volume, text))
    scored.sort(key=lambda x: (-x[0], -x[1]))
    out, seen = [], set()
    for score, hv, text in scored:
        low = text.lower()
        if low in seen:
            continue
        seen.add(low)
        out.append(text)
        if len(out) >= n:
            break
    return out


# ── Build unified rows: one per blog post + one per pillar page ─────────────
rows = []

for post in posts:
    link = link_by_key.get(post['slug'], {})
    cluster = link.get('cluster', 'Uncategorised')
    pillar_url = link.get('pillarPage', '')
    pillar_topic = pillar_title_by_url.get(pillar_url, '')
    secondary = secondary_keywords_for_post(post)
    intent = title_to_intent.get(post['title'], 'Informational') or 'Informational'
    rows.append({
        'type': 'post',
        'primary_keyword': post.get('primaryKeyword') or post['title'],
        'secondary_keywords': '; '.join(secondary),
        'cluster': cluster,
        'pillar_topic': pillar_topic,
        'post_title': post['title'],
        'post_url': link.get('url', f"/news#{post['slug']}"),
        'search_intent': intent,
        'post_status': 'Published',
        'last_updated': post.get('date', ''),
    })

for pillar_url, pillar_link in [(k, v) for k, v in link_by_key.items() if v['type'] == 'pillar']:
    kws = pillar_keywords.get(pillar_url, [])
    primary = kws[0] if kws else pillar_link['title']
    secondary = kws[1:6] if len(kws) > 1 else []
    rows.append({
        'type': 'pillar',
        'primary_keyword': primary,
        'secondary_keywords': '; '.join(secondary),
        'cluster': pillar_link.get('cluster', 'Uncategorised'),
        'pillar_topic': pillar_link['title'],
        'post_title': pillar_link['title'],
        'post_url': pillar_url,
        'search_intent': infer_pillar_intent(pillar_link['title']),
        'post_status': 'Published',
        'last_updated': 'N/A (static page)',
    })

# ── Cluster Status: how well-supported is each cluster? ──────────────────────
cluster_post_counts = defaultdict(int)
for r in rows:
    if r['type'] == 'post':
        cluster_post_counts[r['cluster']] += 1


def cluster_status(cluster):
    n = cluster_post_counts.get(cluster, 0)
    if n >= 5:
        return f'Established ({n} posts)'
    if n >= 2:
        return f'Growing ({n} posts)'
    return f'Needs More Support ({n} post{"s" if n != 1 else ""})'


for r in rows:
    r['cluster_status'] = cluster_status(r['cluster'])

# ── Keyword-cannibalization guard: flag duplicate primary keywords ──────────
pk_owners = defaultdict(list)
for r in rows:
    pk_owners[r['primary_keyword'].strip().lower()].append(r['post_title'])
cannibalization = {k: v for k, v in pk_owners.items() if len(v) > 1}

# ── Write workbook ────────────────────────────────────────────────────────────
wb = Workbook()
ws_out = wb.active
ws_out.title = 'Keyword Cluster Sheet'

HEADERS = [
    'Primary Keyword', 'Secondary Keywords', 'Keyword Cluster', 'Pillar Topic',
    'Associated Post Title', 'Post URL', 'Search Intent', 'Post Status',
    'Cluster Status', 'Last Updated Date',
]
ws_out.append(HEADERS)
for c in range(1, len(HEADERS) + 1):
    cell = ws_out.cell(row=1, column=c)
    cell.font = Font(name='Arial', bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='1F4E78')
    cell.alignment = Alignment(vertical='center', wrap_text=True)
ws_out.freeze_panes = 'A2'
ws_out.row_dimensions[1].height = 32

PILLAR_FILL = PatternFill('solid', start_color='D9E1F2')
CANNIBAL_FILL = PatternFill('solid', start_color='F8CBAD')

rows.sort(key=lambda r: (r['cluster'], r['type'] != 'pillar', r['post_title']))

for r in rows:
    row_idx = ws_out.max_row + 1
    ws_out.append([
        r['primary_keyword'],
        r['secondary_keywords'],
        r['cluster'],
        r['pillar_topic'],
        r['post_title'],
        r['post_url'],
        r['search_intent'],
        r['post_status'],
        r['cluster_status'],
        r['last_updated'],
    ])
    if r['type'] == 'pillar':
        for c in range(1, len(HEADERS) + 1):
            ws_out.cell(row=row_idx, column=c).fill = PILLAR_FILL
    if r['primary_keyword'].strip().lower() in cannibalization:
        ws_out.cell(row=row_idx, column=1).fill = CANNIBAL_FILL

for row in ws_out.iter_rows(min_row=2, max_row=ws_out.max_row, min_col=1, max_col=len(HEADERS)):
    for cell in row:
        cell.font = Font(name='Arial')
        cell.alignment = Alignment(vertical='top', wrap_text=(cell.column in (2,)))

widths = [30, 55, 24, 24, 40, 26, 14, 12, 24, 16]
for i, w in enumerate(widths, start=1):
    ws_out.column_dimensions[get_column_letter(i)].width = w

ws_out.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{ws_out.max_row}'

# ── Summary / Read Me sheet ──────────────────────────────────────────────────
summary = wb.create_sheet('Summary')
summary.append(['Keyword Cluster Sheet — Summary'])
summary.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=13)
summary.append(['Legend: blue-shaded rows = pillar pages. Orange-shaded Primary Keyword = shared with another row (possible cannibalization).'])
summary.append([])

metrics = [
    ('Total blog posts', len([r for r in rows if r['type'] == 'post'])),
    ('Total pillar pages', len([r for r in rows if r['type'] == 'pillar'])),
    ('Total keyword clusters', len(cluster_post_counts)),
    ('Clusters needing more support (<2 posts)', len([c for c, n in cluster_post_counts.items() if n < 2])),
    ('Potential keyword cannibalization (shared primary keyword)', len(cannibalization)),
]
summary.append(['Metric', 'Value'])
for c in (1, 2):
    cell = summary.cell(row=summary.max_row, column=c)
    cell.font = Font(name='Arial', bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='1F4E78')
for label, val in metrics:
    summary.append([label, val])

summary.append([])
summary.append(['Clusters (post count):'])
summary.cell(row=summary.max_row, column=1).font = Font(name='Arial', bold=True)
for cluster, n in sorted(cluster_post_counts.items(), key=lambda x: -x[1]):
    summary.append(['', f'{cluster}: {n} posts — {cluster_status(cluster)}'])

if cannibalization:
    summary.append([])
    summary.append(['Potential keyword cannibalization — same primary keyword used by multiple pages:'])
    summary.cell(row=summary.max_row, column=1).font = Font(name='Arial', bold=True)
    for kw, owners in cannibalization.items():
        summary.append(['', f'"{kw}": {", ".join(owners)}'])

for row in summary.iter_rows():
    for cell in row:
        if cell.row > 3 and cell.column == 1 and not cell.font.bold:
            cell.font = Font(name='Arial')
summary.column_dimensions['A'].width = 45
summary.column_dimensions['B'].width = 75

wb.save('UKMLA_Keyword_Cluster_Sheet.xlsx')
print('Saved UKMLA_Keyword_Cluster_Sheet.xlsx')
print(f"Rows: {len(rows)} ({len([r for r in rows if r['type']=='post'])} posts, {len([r for r in rows if r['type']=='pillar'])} pillars)")
print(f"Clusters: {len(cluster_post_counts)}")
if cannibalization:
    print(f"WARNING: {len(cannibalization)} potential keyword-cannibalization case(s):")
    for kw, owners in cannibalization.items():
        print(f"  \"{kw}\": {', '.join(owners)}")
else:
    print('No keyword cannibalization detected (no duplicate primary keywords).')
