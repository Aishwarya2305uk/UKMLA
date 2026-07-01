import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

with open('scripts/linkage-data.json', encoding='utf-8') as f:
    rows = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = 'Internal Linking Map'

HEADERS = [
    'Post/Page Title', 'URL', 'Type', 'Incoming Links', 'Outgoing Links',
    'Linked From', 'Links To', 'Content Cluster', 'Pillar Page',
]
ws.append(HEADERS)
for c in range(1, len(HEADERS) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(name='Arial', bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='1F4E78')
    cell.alignment = Alignment(vertical='center', wrap_text=True)
ws.freeze_panes = 'A2'
ws.row_dimensions[1].height = 32

key_to_title = {r['key']: r['title'] for r in rows}

def titles(keys):
    return '; '.join(key_to_title.get(k, k) for k in keys) if keys else ''

LOW_FILL = PatternFill('solid', start_color='FFF2CC')  # <2 incoming
ZERO_FILL = PatternFill('solid', start_color='F8CBAD')  # 0 incoming

for r in rows:
    row_idx = ws.max_row + 1
    ws.append([
        r['title'],
        r['url'],
        'Blog Post' if r['type'] == 'post' else 'Pillar Page',
        r['incomingCount'],
        r['outgoingCount'],
        titles(r['linkedFrom']),
        titles(r['linksTo']),
        r['cluster'],
        key_to_title.get(r['pillarPage'], r['pillarPage']),
    ])
    inc_cell = ws.cell(row=row_idx, column=4)
    if r['incomingCount'] == 0:
        inc_cell.fill = ZERO_FILL
    elif r['incomingCount'] < 2:
        inc_cell.fill = LOW_FILL

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(HEADERS)):
    for cell in row:
        cell.font = Font(name='Arial', bold=(cell.row == 1), color='FFFFFF' if cell.row == 1 else '000000')
        if cell.row > 1:
            cell.alignment = Alignment(vertical='top', wrap_text=(cell.column in (6, 7)))

widths = [40, 26, 12, 14, 14, 55, 55, 26, 26]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{ws.max_row}'

# ── Summary sheet ────────────────────────────────────────────────────────
summary = wb.create_sheet('Summary')
summary.append(['Metric', 'Value'])
for c in (1, 2):
    cell = summary.cell(row=1, column=c)
    cell.font = Font(name='Arial', bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='1F4E78')

posts = [r for r in rows if r['type'] == 'post']
pillars = [r for r in rows if r['type'] == 'pillar']
zero_incoming = [r['title'] for r in posts if r['incomingCount'] == 0]
low_incoming = [r['title'] for r in posts if 0 < r['incomingCount'] < 2]
top_incoming = sorted(posts, key=lambda r: -r['incomingCount'])[:10]
top_outgoing = sorted(posts, key=lambda r: -r['outgoingCount'])[:10]

metrics = [
    ('Total blog posts', len(posts)),
    ('Total pillar/site pages tracked', len(pillars)),
    ('Total outgoing links (from posts)', sum(r['outgoingCount'] for r in posts)),
    ('Avg outgoing links per post', round(sum(r['outgoingCount'] for r in posts) / len(posts), 1) if posts else 0),
    ('Posts with 0 incoming links', len(zero_incoming)),
    ('Posts with 1 incoming link', len(low_incoming)),
]
for label, val in metrics:
    summary.append([label, val])

summary.append([])
summary.append(['Posts needing more internal links (0 incoming):'])
summary.cell(row=summary.max_row, column=1).font = Font(name='Arial', bold=True)
for t in zero_incoming:
    summary.append(['', t])

summary.append([])
summary.append(['Most heavily interlinked posts (by incoming links):'])
summary.cell(row=summary.max_row, column=1).font = Font(name='Arial', bold=True)
for r in top_incoming:
    summary.append(['', f"{r['title']} ({r['incomingCount']} incoming)"])

for row in summary.iter_rows():
    for cell in row:
        if cell.row > 1:
            cell.font = Font(name='Arial')
summary.column_dimensions['A'].width = 40
summary.column_dimensions['B'].width = 60

wb.save('UKMLA_Internal_Linking_Map.xlsx')
print('Saved UKMLA_Internal_Linking_Map.xlsx')
print('Posts:', len(posts), 'Pillars:', len(pillars))
print('Zero-incoming posts:', len(zero_incoming))
